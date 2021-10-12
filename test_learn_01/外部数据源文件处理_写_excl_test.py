from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# 文本地址
# https://openpyxl.readthedocs.io/en/stable/usage.html
wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
 ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
       _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

ws4 =  wb.create_sheet(title="刘基级")
for i in range(1,31):
   ws4.cell(column=1 , row=i, value="垃圾")


print(ws3['AA10'].value)


wb.save(filename = dest_filename)